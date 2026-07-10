from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None


@dataclass(frozen=True)
class TableSummary:
    file: Path
    sheet: str
    rows: int
    columns: int
    empty_cells: int
    headers: list[str]


def summarize_tables(input_dir: Path, output_csv: Path) -> list[TableSummary]:
    if not input_dir.exists():
        raise FileNotFoundError(f"Table folder does not exist: {input_dir}")

    summaries: list[TableSummary] = []
    for path in sorted(input_dir.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() == ".csv":
            summaries.append(summarize_csv(path))
        elif path.suffix.lower() == ".xlsx":
            summaries.extend(summarize_xlsx(path))

    write_summary_csv(summaries, output_csv)
    return summaries


def summarize_csv(path: Path) -> TableSummary:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    return summarize_rows(path, "csv", rows)


def summarize_xlsx(path: Path) -> list[TableSummary]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to summarize .xlsx files")

    workbook = load_workbook(path, read_only=True, data_only=True)
    summaries = []
    for sheet in workbook.worksheets:
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        summaries.append(summarize_rows(path, sheet.title, rows))
    workbook.close()
    return summaries


def summarize_rows(path: Path, sheet: str, rows: list[list[object]]) -> TableSummary:
    if not rows:
        return TableSummary(path, sheet, 0, 0, 0, [])

    max_columns = max(len(row) for row in rows)
    normalized = [row + [""] * (max_columns - len(row)) for row in rows]
    headers = [str(value).strip() for value in normalized[0]]
    data_rows = normalized[1:]
    empty_cells = sum(1 for row in data_rows for value in row if value in (None, ""))

    return TableSummary(
        file=path,
        sheet=sheet,
        rows=len(data_rows),
        columns=max_columns,
        empty_cells=empty_cells,
        headers=headers,
    )


def write_summary_csv(summaries: Iterable[TableSummary], output_csv: Path) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["file", "sheet", "rows", "columns", "empty_cells", "headers"])
        for summary in summaries:
            writer.writerow(
                [
                    summary.file.name,
                    summary.sheet,
                    summary.rows,
                    summary.columns,
                    summary.empty_cells,
                    " | ".join(summary.headers),
                ]
            )
